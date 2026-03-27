import matplotlib.pyplot as plt
import openpyxl


def plot_sheets_1(workbook_path, output_folder):
    # Load the workbook
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    fig, ax = plt.subplots(1, 3, figsize=(15, 5))
    fig2, ax2 = plt.subplots(1, 3, figsize=(15, 5))
    colors = ['tab:orange', 'tab:blue', 'tab:green']
    for i, sheet_name in enumerate(wb.sheetnames[-3:]):
        sheet = wb[sheet_name]
        data = list(sheet.iter_rows(values_only=True))

        # Extract columns
        x_values = [row[0] for row in data if row[0] is not None][2:]  # Column 1
        E_implicit = [row[1] for row in data if row[1] is not None][2:] # Column 2
        E_explicit = [row[2] for row in data if row[2] is not None][2:]  # Column 3
        P_CO2_implicit = [row[3] for row in data if row[3] is not None][2:] # Column 2
        P_CO2_explicit = [row[4] for row in data if row[4] is not None][2:]  # Column 3
        P_equil_implicit = [row[5] for row in data if row[5] is not None][2:] # Column 2
        P_equil_explicit = [row[6] for row in data if row[6] is not None][2:]  # Column 3

        Tl, alpha, L_G = sheet_name.split(",")
        Tl = Tl[3:]
        alpha = alpha[6:]
        L_G = L_G[5:]

        title = "$T_{L}$=" + Tl + ", " + r"$\alpha$=" + alpha + ", " + r"$\frac{L}{G}$=" + L_G + "\n" + f"CO2 % Org|New: {data[1][7]:.3f}% | {data[1][8]:.3f}% "

        # Create the plot
        ax[i].plot(x_values, E_implicit, linestyle='-', color=colors[i], label="Implicit Enhancement Factor")
        ax[i].plot(x_values, E_explicit, linestyle='--', color=colors[i], label="Explicit Enhancement Factor")
        ax[i].set_ylabel("Enhancement Factor")
        ax[i].set_xlabel("Column Height (m)",)
        ax[i].set_title(title)
        ax[i].legend(loc='best')

        # Create the plot
        ax2[i].plot(x_values, P_CO2_implicit, linestyle='-', color='tab:blue', label="Implicit CO2 Vapor Pressure")
        ax2[i].plot(x_values, P_CO2_explicit, linestyle='--', color='tab:blue', label="Explicit CO2 Vapor Pressure")
        ax2[i].plot(x_values, P_equil_implicit, linestyle='-', color='tab:green', label="Implicit Equilibrium Pressure")
        ax2[i].plot(x_values, P_equil_explicit, linestyle='--', color='tab:green', label="Explicit Equilibrium Pressure")
        ax2[i].set_ylabel("Pressure (Pa)")
        ax2[i].set_xlabel("Column Height (m)",)
        ax2[i].set_title(title)

    # Save the plot
    # plt.savefig(f"{output_folder}/{sheet_name}.png", bbox_inches="tight")
    plt.rc('axes', titlesize=17)  # Font size for titles
    plt.rc('axes', labelsize=16)  # Font size for axis labels
    plt.rc('xtick', labelsize=14)  # Font size for x-axis tick labels
    plt.rc('ytick', labelsize=14)  # Font size for y-axis tick labels
    plt.rc('lines', linewidth=3)  # Font size for y-axis tick labels
    # plt.subplots_adjust(left=0.0, right=1.0, top=0.9, bottom=0.1, wspace=0.5, hspace=0.5)
    plt.show()
    plt.savefig(r"plot_comparison.png", bbox_inches='tight', dpi=500, pad_inches=0.2)
    plt.close()


# Example usage:
# combine_workbooks("workbook1.xlsx", "workbook2.xlsx", "combined_workbook.xlsx")
# reorder_columns("combined_workbook.xlsx", "reordered_workbook.xlsx")
plot_sheets_1(r"combined_low_capture.xlsx",
              r"output_plots_low_capture")
